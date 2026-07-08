"""
importador_usuarios.py - Utilidad para importar usuarios desde Excel y CSV.
"""

import csv
import io
from typing import Any, Dict, List, Optional

from werkzeug.datastructures import FileStorage


class ImportadorUsuarios:
    """Importa usuarios desde archivos CSV y XLSX."""

    def __init__(self, fachada: Any) -> None:
        """Inicializa el importador con acceso a la fachada."""
        self._fachada = fachada

    def procesar_archivo(self, archivo: FileStorage) -> Dict[str, Any]:
        """
        Procesa un archivo de importación (CSV o XLSX).
        
        Retorna un diccionario con estadísticas de importación.
        """
        if archivo.filename.endswith(".csv"):
            return self._procesar_csv(archivo)
        elif archivo.filename.endswith(".xlsx"):
            return self._procesar_xlsx(archivo)
        else:
            raise ValueError("Formato de archivo no soportado.")

    def _procesar_csv(self, archivo: FileStorage) -> Dict[str, Any]:
        """Procesa un archivo CSV."""
        exitosos = 0
        errores = 0
        errores_detalle = []

        try:
            # Leer archivo CSV
            contenido = archivo.read().decode("utf-8")
            lector = csv.DictReader(io.StringIO(contenido))

            for fila_num, fila in enumerate(lector, start=2):  # Empieza en 2 (header es 1)
                try:
                    usuario = self._procesar_fila(fila)
                    if usuario:
                        self._fachada.crear_usuario(usuario)
                        exitosos += 1
                except ValueError as e:
                    errores += 1
                    errores_detalle.append(f"Fila {fila_num}: {str(e)}")
                except Exception as e:
                    errores += 1
                    errores_detalle.append(f"Fila {fila_num}: Error inesperado: {str(e)}")

        except Exception as e:
            raise ValueError(f"Error al leer archivo CSV: {str(e)}")

        return {
            "exitosos": exitosos,
            "errores": errores,
            "errores_detalle": errores_detalle[:10],  # Mostrar solo primeros 10
        }

    def _procesar_xlsx(self, archivo: FileStorage) -> Dict[str, Any]:
        """Procesa un archivo XLSX."""
        try:
            import openpyxl
        except ImportError:
            raise ValueError("openpyxl no está instalado. Instálalo con: pip install openpyxl")

        exitosos = 0
        errores = 0
        errores_detalle = []

        try:
            # Cargar workbook
            contenido = archivo.read()
            workbook = openpyxl.load_workbook(io.BytesIO(contenido))
            hoja = workbook.active

            # Obtener encabezados
            encabezados = [cell.value for cell in hoja[1]]

            for fila_num, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Crear diccionario de datos
                    datos_fila = dict(zip(encabezados, fila))
                    usuario = self._procesar_fila(datos_fila)
                    if usuario:
                        self._fachada.crear_usuario(usuario)
                        exitosos += 1
                except ValueError as e:
                    errores += 1
                    errores_detalle.append(f"Fila {fila_num}: {str(e)}")
                except Exception as e:
                    errores += 1
                    errores_detalle.append(f"Fila {fila_num}: Error inesperado: {str(e)}")

        except ImportError:
            raise ValueError("Debes instalar openpyxl: pip install openpyxl")
        except Exception as e:
            raise ValueError(f"Error al leer archivo XLSX: {str(e)}")

        return {
            "exitosos": exitosos,
            "errores": errores,
            "errores_detalle": errores_detalle[:10],
        }

    @staticmethod
    def _procesar_fila(fila: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """
        Procesa una fila y retorna un diccionario de usuario validado.
        
        Campos esperados: nombre, apellido, documento, email, telefono, rol
        """
        # Limpiar valores
        fila = {k: (v.strip() if isinstance(v, str) else v) for k, v in fila.items() if v}

        # Validar campos requeridos
        campos_requeridos = ["nombre", "apellido", "documento", "email", "rol"]
        for campo in campos_requeridos:
            if campo not in fila or not fila[campo]:
                raise ValueError(f"Campo requerido '{campo}' ausente o vacío.")

        # Validar rol
        roles_validos = ["administrador", "docente", "estudiante", "coordinador"]
        if fila["rol"] not in roles_validos:
            raise ValueError(f"Rol '{fila['rol']}' no es válido. Válidos: {', '.join(roles_validos)}")

        # Validar email
        if "@" not in fila["email"]:
            raise ValueError(f"Email '{fila['email']}' no es válido.")

        return {
            "nombre": fila["nombre"],
            "apellido": fila["apellido"],
            "documento": str(fila["documento"]).strip(),
            "email": fila["email"],
            "telefono": fila.get("telefono", ""),
            "rol": fila["rol"],
        }

    @staticmethod
    def generar_template_csv() -> str:
        """Genera un template CSV para importación."""
        return """nombre,apellido,documento,email,telefono,rol
Juan,Pérez,1234567890,juan@unilevel.edu,5551234567,estudiante
María,González,9876543210,maria@unilevel.edu,5559876543,docente
Carlos,López,5555555555,carlos@unilevel.edu,5551111111,coordinador"""

    @staticmethod
    def generar_template_xlsx() -> bytes:
        """Genera un template XLSX para importación."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ValueError("openpyxl no está instalado.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios"

        # Encabezados
        encabezados = ["nombre", "apellido", "documento", "email", "telefono", "rol"]
        ws.append(encabezados)

        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Datos de ejemplo
        ejemplos = [
            ["Juan", "Pérez", "1234567890", "juan@unilevel.edu", "5551234567", "estudiante"],
            ["María", "González", "9876543210", "maria@unilevel.edu", "5559876543", "docente"],
            ["Carlos", "López", "5555555555", "carlos@unilevel.edu", "5551111111", "coordinador"],
        ]

        for fila in ejemplos:
            ws.append(fila)

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # Guardar a bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

        wb.save(output)
        output.seek(0)
        return output.getvalue()
